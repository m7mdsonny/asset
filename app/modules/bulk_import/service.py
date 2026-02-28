"""Bulk import from Excel: employees and assets."""

from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.employee import Employee


class BulkImportService:
    def __init__(self, db: AsyncSession) -> None:
        self._db = db

    async def import_employees(
        self,
        company_id: UUID,
        branch_id: UUID,
        file_path: str,
    ) -> tuple[int, list[str]]:
        """Import employees from Excel. Sheet 'Employees', columns: name, national_id, job_title, department."""
        wb = load_workbook(file_path, read_only=True)
        ws = wb["Employees"] if "Employees" in wb.sheetnames else wb.active
        created = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            national_id = str(row[1]).strip() if len(row) > 1 and row[1] else None
            job_title = str(row[2]).strip() if len(row) > 2 and row[2] else None
            department = str(row[3]).strip() if len(row) > 3 and row[3] else None
            try:
                emp = Employee(
                    company_id=company_id,
                    branch_id=branch_id,
                    name=name,
                    national_id=national_id or None,
                    job_title=job_title or None,
                    department=department or None,
                    status="active",
                )
                self._db.add(emp)
                created += 1
            except Exception as e:
                errors.append(f"Row {name}: {e}")
        await self._db.flush()
        return created, errors

    async def import_assets(
        self,
        company_id: UUID,
        branch_id: UUID,
        file_path: str,
    ) -> tuple[int, list[str]]:
        """Import assets from Excel. Sheet 'Assets', columns: type, brand, model, serial_number, ..."""
        wb = load_workbook(file_path, read_only=True)
        ws = wb["Assets"] if "Assets" in wb.sheetnames else wb.active
        created = 0
        errors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            asset_type = str(row[0]).strip().lower() or "other"
            if asset_type not in ("laptop", "mobile", "tablet", "desktop", "monitor", "peripheral", "other"):
                asset_type = "other"
            brand = str(row[1]).strip() if len(row) > 1 and row[1] else None
            model = str(row[2]).strip() if len(row) > 2 and row[2] else None
            serial_number = str(row[3]).strip() if len(row) > 3 and row[3] else None
            try:
                asset = Asset(
                    company_id=company_id,
                    branch_id=branch_id,
                    type=asset_type,
                    brand=brand or None,
                    model=model or None,
                    serial_number=serial_number or None,
                    status="active",
                )
                self._db.add(asset)
                created += 1
            except Exception as e:
                errors.append(f"Row {serial_number or model or '?'}: {e}")
        await self._db.flush()
        return created, errors
